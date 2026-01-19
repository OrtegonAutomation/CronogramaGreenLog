"""
CRONOGRAMA DE PROYECTO GREENLOG - CENIT
Herramienta para gestión de cronograma, registro de avances y generación de reportes

Características:
- Genera Excel con cronograma estructurado
- Crea imagen del diagrama de Gantt
- Permite registrar tiempos y reportar avances
- Exporta reportes de estado

Uso:
    python cronograma_cenit.py [comando]
    
    Comandos:
        crear       - Crea el cronograma inicial (Excel + imagen Gantt)
        avance      - Registra avance de una actividad
        reporte     - Genera reporte de estado actual
        actualizar  - Actualiza la imagen del Gantt con los avances
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from datetime import datetime, timedelta
import os
import json
import sys

# ============================================================================
# CONFIGURACIÓN DEL PROYECTO
# ============================================================================
FECHA_INICIO = datetime(2025, 1, 20)
ARCHIVO_JSON = "cronograma_data.json"
ARCHIVO_JS = "cronograma_data.js"
ARCHIVO_EXCEL = "Cronograma_CENIT_COMPLETO.xlsx"
ARCHIVO_IMAGEN = "Cronograma_Gantt_CENIT.png"

def cargar_datos_json():
    """Carga los datos desde el archivo JSON central"""
    if os.path.exists(ARCHIVO_JSON):
        with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    print(f"[ERROR] No se encontró {ARCHIVO_JSON}")
    return None

def sincronizar_datos():
    """Sincroniza JSON a JS para evitar errores de CORS"""
    data_full = cargar_datos_json()
    if not data_full: return
    
    js_content = f"const CRONOGRAMA_DATA = {json.dumps(data_full, indent=2, ensure_ascii=False)};"
    with open(ARCHIVO_JS, 'w', encoding='utf-8') as f:
        f.write(js_content)
    print(f"[OK] Datos sincronizados para Web: {ARCHIVO_JS}")

def crear_excel_completo():
    """Genera Excel con todas las actividades y un Gantt Visual usando colores"""
    data_full = cargar_datos_json()
    if not data_full: return
    
    config = data_full.get("config", {})
    actividades = data_full.get("actividades", [])
    
    # Obtener fecha de inicio del JSON o usar default
    fecha_base_str = config.get("fecha_inicio", "2026-01-20")
    fecha_base = datetime.strptime(fecha_base_str, "%Y-%m-%d")

    filas = []
    fase_actual = ""
    
    for item in actividades:
        if item["tipo"] == "PRINCIPAL":
            fase_actual = item.get("fase", "")
            
        semana = item["semana_inicio"]
        duracion = item["duracion"]
        inicio = fecha_base + timedelta(weeks=semana - 1)
        fin = inicio + timedelta(weeks=duracion) - timedelta(days=3)
        
        filas.append({
            "Fase": fase_actual,
            "Actividad": item["actividad"],
            "Tipo": "Principal" if item["tipo"] == "PRINCIPAL" else "Sub-actividad",
            "Inicio": inicio.strftime("%d/%m/%Y"),
            "Fin": fin.strftime("%d/%m/%Y"),
            "S. Inicio": semana,
            "Dur": duracion,
            "Resp": item["responsable"],
            "Avance": f"{item['porcentaje']}%",
            "Estado": item["estado"]
        })

    df = pd.DataFrame(filas)
    
    with pd.ExcelWriter(ARCHIVO_EXCEL, engine='openpyxl') as writer:
        # 1. Hoja de Datos
        df.to_excel(writer, sheet_name='Detalle', index=False)
        
        # 2. Hoja de Gantt Visual
        workbook = writer.book
        worksheet = workbook.create_sheet('Gantt Visual')
        
        # Estilos
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill(start_color="004a99", end_color="004a99", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Cabeceras del Gantt
        worksheet.cell(row=1, column=1, value="Actividad").font = Font(bold=True)
        worksheet.column_dimensions['A'].width = 50
        
        for w in range(1, 21):
            cell = worksheet.cell(row=1, column=w+1, value=f"S{w}")
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center')
        
        # Dibujar Actividades y Barras
        for i, item in enumerate(actividades):
            row_idx = i + 2
            # Nombre de actividad
            cell_name = worksheet.cell(row=row_idx, column=1, value=item["actividad"])
            if item["tipo"] == "PRINCIPAL":
                cell_name.font = Font(bold=True)
                cell_name.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Barras de tiempo
            s_ini = item["semana_inicio"]
            dur = item["duracion"]
            color_hex = item.get("color", "#38bdf8").replace("#", "")
            if item["tipo"] == "SUB": color_hex = "475569"
            
            bar_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            
            for w in range(s_ini, s_ini + dur):
                if w <= 20:
                    cell_bar = worksheet.cell(row=row_idx, column=w+1)
                    cell_bar.fill = bar_fill
                    cell_bar.border = thin_border
                    if item["porcentaje"] == 100:
                        cell_bar.value = "X"
                        cell_bar.alignment = Alignment(horizontal='center')
                        cell_bar.font = Font(color="FFFFFF", bold=True)

    print(f"[OK] Excel con Gantt Visual generado: {ARCHIVO_EXCEL}")

def crear_gantt_profesional():
    """Genera imagen Gantt mejorada con nombres a la izquierda"""
    data_full = cargar_datos_json()
    if not data_full: return

    config = data_full.get("config", {})
    actividades = data_full.get("actividades", [])
    
    # Filtrar solo principales para la imagen (para que sea legible)
    principales = [item for item in actividades if item["tipo"] == "PRINCIPAL"]
    principales.reverse() # Para que la primera aparezca arriba
    
    fig, ax = plt.subplots(figsize=(14, 8))
    
    y_pos = range(len(principales))
    nombres = [item["actividad"] for item in principales]
    
    for i, item in enumerate(principales):
        inicio = item["semana_inicio"]
        duracion = item["duracion"]
        color = item.get("color", "#3498db")
        
        ax.barh(i, duracion, left=inicio, height=0.6, color=color, alpha=0.8, edgecolor='black')
        if item["porcentaje"] > 0:
            ax.barh(i, duracion * (item["porcentaje"]/100), left=inicio, height=0.3, color='white', alpha=0.5)

    ax.set_yticks(y_pos)
    ax.set_yticklabels(nombres, fontsize=10)
    ax.set_xlabel('Semanas del Proyecto')
    ax.set_title('PROYECTO GREENLOG - CENIT (Gantt Ejecutivo)')
    ax.grid(axis='x', linestyle='--', alpha=0.7)
    
    # Marcador de hoy dinámico
    fecha_base_str = config.get("fecha_inicio", "2026-01-20")
    start_date = datetime.strptime(fecha_base_str, "%Y-%m-%d")
    days_since = (datetime.now() - start_date).days
    weeks_since = days_since / 7 + 1
    
    if 1 <= weeks_since <= 20:
        ax.axvline(x=weeks_since, color='red', linestyle='-', linewidth=2, label='Hoy')
        ax.legend()

    plt.tight_layout()
    plt.savefig(ARCHIVO_IMAGEN, dpi=150)
    print(f"[OK] Imagen Gantt ejecutiva creada: {ARCHIVO_IMAGEN}")

def main():
    if len(sys.argv) < 2:
        print("\nComandos:")
        print("  python cronograma_cenit.py sync    - Sincroniza JSON, JS, Excel y Gantt")
        print("  python cronograma_cenit.py open    - Abre el Dashboard")
        return
        
    cmd = sys.argv[1].lower()
    if cmd == "sync":
        sincronizar_datos()
        crear_excel_completo()
        crear_gantt_profesional()
    elif cmd == "open":
        print("Abriendo Dashboard...")
        os.startfile("index.html")

if __name__ == "__main__":
    main()
